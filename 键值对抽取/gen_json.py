import os
import json
import re
from io import StringIO

from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
import os
import json
import re
from io import StringIO
import xlrd
from py2neo import Graph, Node, Relationship
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


PDF_DIR = '0601/测试文档'
JSON_FILE_PATH = '0601/输出结果/输出pdf抽取的信息_基于规则.json'

train_json_path='329/t/3.json'
own_json_path='newtrain.json'
val_dir='val'
train_dir='train'
val_json_path='329/v/验证集.json'

modeljson='329/v/输出pdf抽取的信息_使用模型.json'
rulejson='329/v/输出pdf抽取的信息_基于规则.json'
truejson='329/v/验证集.json'
resultexcel='329/v/实验结果.xlsx'
list1=['姓名', '出生年月', '性别', '电话', '籍贯', '落户市县', '政治面貌', '学位', '毕业时间', '工作时间', '项目时间', '毕业院校', '工作单位',
                 '工作内容', '职务', '项目名称', '项目责任']
listt=['毕业院校', '工作单位','工作内容', '职务', '项目名称', '项目责任']#不能改
listt1=[ '学位', '毕业时间', '工作时间', '项目时间']
listtt1=['姓名', '出生年月', '性别', '电话', '籍贯', '落户市县', '政治面貌', '学位', '毕业时间', '工作时间', '项目时间', '毕业院校', '工作单位',
                 '工作内容', '职务', '项目名称', '项目责任']
listt2=['毕业院校', '工作单位','工作内容', '职务', '项目名称', '项目责任']

def build_train_val_json_by_own_train_json(own_json_path, val_dir,  val_json_path):
    with open(own_json_path, 'r') as j:
        own_filename_to_info = json.load(j)
    train_filename_to_info = {}
    val_filename_to_info = {}
    val_paths = os.listdir(val_dir)
    for p in val_paths:
        if p.endswith('.pdf'):
            val_filename = p[:-4]
            val_filename_to_info[val_filename] = own_filename_to_info[val_filename]
    with open(val_json_path, 'w') as j:
        json.dump(val_filename_to_info, j)
    print('finish')

def counttrain(train_dir,train_json_path,key):
    typeflag = 3;
    count = 0
    rrcount = 0
    mmcount = 0
    rcount = 0
    mcount = 0
    falsecount = 0
    pr = 0
    rr = 0
    pm = 0
    rm = 0
    num = 0

    with open(train_json_path,'r', encoding='utf-8') as j:
        truejson_info = json.load(j)
        val_paths = os.listdir(train_dir)
        for p in val_paths:
            if p.endswith('.pdf'):
                val_filename = p[:-4]
                num = num + 1
                info1 = truejson_info[val_filename]
                try:
                    if isinstance(info1[key], str):
                        typeflag = 0
                        if key in info1:
                            count += 1

                    if isinstance(info1[key], list):
                        typeflag = 1
                        if key in info1:
                            count += len(info1[key])
                except Exception as e:
                    f = 1

    print(count)
    return count

def allcount(list1):
    listpr=[]
    listrr=[]
    listpm=[]
    listrm=[]
    listw=[]
    listf1r=[]
    listf1m = []
    prweighted = 0
    rrweighted = 0
    pmweighted = 0
    rmweighted = 0
    allw=0
    for i in list1:
        key = i
        counttrain(val_dir,val_json_path,key)


def scoretest(modeljson,rulejson,truejson,val_dir,key):
    typeflag=3;
    count=0
    rrcount=0
    mmcount=0
    rcount=0
    mcount=0
    falsecount=0
    pr=0
    rr=0
    pm=0
    rm=0
    num=0
    with open(modeljson, 'r',encoding='utf-8') as j:
        modeljson_info = json.load(j)
    with open(rulejson, 'r',encoding='utf-8') as j:
        rulejson_info = json.load(j)
    with open(truejson, 'r',encoding='utf-8') as j:
        truejson_info = json.load(j)
        val_paths = os.listdir(val_dir)
        for p in val_paths:
            if p.endswith('.pdf'):
                val_filename = p[:-4]
                num=num+1
                info3=modeljson_info[val_filename]
                info2=rulejson_info[val_filename]
                info1=truejson_info[val_filename]

            try:
                if isinstance (info1[key],str):
                    typeflag = 0
                    if key in info1:
                        count += 1
                    if key in info2:
                        rrcount += 1
                    if key in info3:
                        mmcount += 1
                    if info1[key] == info2[key]:
                        rcount += 1

                        # print(key,'规则',rcount,'T    ',info1[key])
                    # else:print(key,'规则',rcount,'F    ',info2[key],'          规则错，正确是:',info1[key],'                 原始文档名：',val_filename)
                    if info1[key] == info3[key]:
                        mcount += 1

                        # print(key,'模型',mcount,'T    ',info3[key])
                    # else:print(key,'模型', mcount, 'F    ', info3[key],'          模型错，正确是:',info1[key],'                 原始文档名：',val_filename)
                if isinstance (info1[key],list):
                    typeflag = 1
                    if key in listt:
                        typeflag = 2
                        if key in info1:
                            count += len(info1[key])
                        if key in info2:
                            rrcount += len(info2[key])
                        if key in info3:
                            mmcount +=  len(info3[key])
                        for i in info2[key]:
                            for j in info1[key]:
                               if i in j:
                                    rcount +=len(i)/len(j)

                        for i in info3[key]:
                            for j in info1[key]:
                                if i in j:
                                    mcount += len(i)/len(j)

                    else:
                        if key in info1:
                            count += len(info1[key])
                        if key in info2:
                            rrcount += len(info2[key])
                        if key in info3:
                            mmcount += len(info3[key])
                        for i in info2[key]:
                            if i in info1[key]:
                                rcount +=1
                        for i in info3[key]:
                            if i in info1[key]:
                                mcount += 1
            except Exception as e:f=1

    if typeflag==0:
        print("_________普通字段_________")
    if typeflag==1:
        print("_________列表字段_________")
    if typeflag==2:
        print("_________字符级列表字段_________")
    print('验证集  规则  模型  正确个数：',count,rcount,mcount)
    print('验证集  规则  模型  结果个数：',count, rrcount, mmcount)
    #print('\n')

    w=0
    w = count / num
    if rrcount != 0:
        pr = rcount / rrcount
    if count != 0:
        rr = rcount / count
    if mmcount != 0:
        pm = mcount / mmcount
    if count != 0:
        rm = mcount / count
    f1r = 0
    f1m=0
    if pr+rr!=0 :
        f1r=2*pr*rr/(pr+rr)

    print(key,'\n',w,'\n%.8f'%pr,'\n%.8f'%rr,'\n%.8f'%f1r)
    if pm+rm!=0:
        f1m=2*pm*rm/(pm+rm)

    print('\n',key,'\n',w,'\n%.8f'%pm,'\n%.8f'%rm ,'\n%.8f'%f1m)
    print('\n')
    return pr,rr,f1r,pm,rm,f1m,w



def allscore(list1):
    listpr=[]
    listrr=[]
    listpm=[]
    listrm=[]
    listw=[]
    listf1r=[]
    listf1m = []
    prweighted = 0
    rrweighted = 0
    pmweighted = 0
    rmweighted = 0
    allw=0
    for i in list1:
        key = i
        pr,rr,f1r,pm,rm,f1m,w=scoretest(modeljson, rulejson, truejson, val_dir, key)
        listpr.append(pr)
        listrr.append(rr)
        listpm.append(pm)
        listrm.append(rm)
        listf1r.append(f1r)
        listf1m.append(f1m)
        listw.append(w)
        prweighted+=  pr * w
        rrweighted+=  rr * w
        pmweighted += pm * w
        rmweighted += rm * w
        allw+=w

    prw=  prweighted/ allw
    rrw = rrweighted / allw
    pmw = pmweighted / allw
    rmw = rmweighted / allw
    mf1r=2* prweighted*rrweighted/(prweighted+rrweighted)/allw
    mf1m = 2 * pmweighted * rmweighted / (pmweighted + rmweighted)/allw
    print(mf1r,prw,rrw)
    print(mf1m, pmw, rmw)


    result_wb = Workbook()
    # 第一个sheet是ws
    ws1 = result_wb.worksheets[0]
    # ws1=wb1.create_sheet('result',0)
    # 设置ws的名称
    ws1.title = "简历"

    ft = Font(name='Arial', size=11, bold=True)
    for k in range(len(list1)):
        ws1.cell(row=1, column=k + 2).value = list1[k]
        ws1.cell(row=1, column=k + 2).font = ft
        val_paths = os.listdir(val_dir)
    j = 2
    for p in listpr:
        ws1.cell(row=2, column=1).value='p-规则'
        ws1.cell(row=2, column=j).value = float(p)
        j = j + 1
    j = 2
    for p in listpm:
        ws1.cell(row=3, column=1).value='p-模型'
        ws1.cell(row=3, column=j).value = float(p)
        j = j + 1
    j = 2
    for p in listrr:
        ws1.cell(row=4, column=1).value='r-规则'
        ws1.cell(row=4, column=j).value =float(p)
        j = j + 1
    j = 2
    for p in listrm:
        ws1.cell(row=5, column=1).value='r-模型'
        ws1.cell(row=5, column=j).value = float(p)
        j = j + 1
    j = 2
    for p in listf1r:
        ws1.cell(row=6, column=1).value = 'f1-规则'
        ws1.cell(row=6, column=j).value = float(p)
        j = j + 1
    j = 2
    for p in listf1m:
        ws1.cell(row=7, column=1).value = 'f1-模型'
        ws1.cell(row=7, column=j).value = float(p)
        j = j + 1
    j = 2
    for p in listw:
        ws1.cell(row=8, column=1).value='权重'
        ws1.cell(row=8, column=j).value = float(p/allw)
        j = j + 1


    ws1.cell(row=9, column=1).value = '规则f1'
    ws1.cell(row=10, column=1).value = '模型f1'
    ws1.cell(row=9, column=2).value = float(mf1r)
    ws1.cell(row=10, column=2).value = float(mf1m)
    ws1.cell(row=11, column=1).value = '规则p'
    ws1.cell(row=12, column=1).value = '模型p'
    ws1.cell(row=11, column=2).value = float(prw)
    ws1.cell(row=12, column=2).value = float(pmw)
    ws1.cell(row=13, column=1).value = '规则r'
    ws1.cell(row=14, column=1).value = '模型r'
    ws1.cell(row=13, column=2).value = float(rrw)
    ws1.cell(row=14, column=2).value = float(rmw)


    result_wb.save(filename='结果513.xlsx')








def pdf2strlist(pdf_path):
    strlist = []
    content = ''
    if pdf_path.endswith('.pdf'):
        rsrcmgr = PDFResourceManager(caching=True)
        laparams = LAParams()
        retstr = StringIO()
        device = TextConverter(rsrcmgr, retstr, laparams=laparams)
        with open(pdf_path, 'rb') as fp:
            interpreter = PDFPageInterpreter(rsrcmgr, device)
            for page in PDFPage.get_pages(fp, pagenos=set()):
                page.rotate = page.rotate % 360
                interpreter.process_page(page)
        device.close()
        content = retstr.getvalue()
    if content != '':
        words = content.strip().replace('\n', '').split()
        for word in words:
            word = re.split('[:：]', word)
            for w in word:
                strlist.append(w)
        strlist = list(filter(lambda x: x, strlist))
    return strlist


def cleanstr(s):  # 取右边第一个元素作为value
    info = []
    string1 = []
    lists = list(s)
    for i, e in enumerate(lists):
        if e.isdigit() == False and e != '年' and e != '月' and e != '-' and e != '.'and e != '、':
            info.append(e)
    string1 = ''.join(info)
    return string1

def write_info(content):
    info = {}
    list1=[]
    listname=[]
    list2=[]
    max = 0
    flag = 0
    info['姓名'] = ext_general_field(content, '姓名')#取下一个元素作为value
    info['籍贯'] = ext_general_field(content, '籍贯')
    info['出生年月'] = ext_general_field(content, '出生')
    info['电话'] = ext_general_field(content, '电话')
    info['落户市县'] = ext_general_field(content, '户口')
    info['政治面貌'] = ext_general_field(content, '面貌')
    info['最高学历'] =''
    info['毕业院校']=''
    info['性别'] = ''
    info['学位'] = ''
    info['年龄']=ext_general_field(content, '年龄')
    info['毕业时间'] =''
    info['邮箱'] =  ext_general_field(content, '邮箱')
    info['工作单位'] = ''
    info['工作内容'] = ''
    info['项目经历'] = ''
    info['项目名称'] = ''
    info['项目时间'] = ''
    info['项目责任'] = ''
    info['职务'] =''
    #info['原文全文']=content



    if ('省' or '市' or '县'or '区') not in info['籍贯']:
        info['籍贯']=''

    if info['落户市县']=='':
        info['落户市县'] = ext_general_field(content, '户籍')
    if ('省' or '市' or '县'or '区') not in info['落户市县']:
        info['落户市县']=''

    if info['出生年月'][:2].isdigit() == False:
        info['出生年月']=''

    if info['电话'].isdigit()==0 or len(info['电话'])<7:
        info['电话'] = ''






    try:
        if info['姓名'] == '':
            if len(content[0])<=3:
                listname.append(content[0])
                info['姓名'] = ''.join(listname)

        if len(info['姓名'])==1:
            if len(content[1])==1:
                listname.append(content[1])
                info['姓名'] = ''.join(listname)
                if len(content[2]) == 1 and content[2] != '·':
                    listname.append(content[2])
                    info['姓名']=''.join(listname)

        if ''.join(listname)=='姓名':
            info['姓名']=content[2]




        if info['性别'] == '':
            for c in content:
                if '男'in c:
                    info['性别'] = '男'
                if '女'in c:
                    info['性别'] = '女'

        if info['邮箱'] == '':
            for c in content:
                if '@'in c:
                    info['邮箱'] =c
                if '.com'in c:
                    info['邮箱'] = c

        if info['年龄'] != '':
            if info['年龄'].isdigit()==-1:
                    info['年龄'] = ''

        if info['籍贯'] == '':
            for c in content:
                if '省' in c:
                    info['籍贯'] = c
                if '市' == c[-1:]:
                    info['籍贯'] = c
                    break

        if info['电话'] == '':
            for c in content:
                if c.isdigit():
                    if len(c) == 11 or len(c) == 7:
                        info['电话'] = c



        if info['最高学历'] == '':
            max=0
            flag=0
            for c in content:
                if '小学' in c:
                    flag=0
                if '初中'in c:
                    flag=1
                if '高中'in c:
                    flag = 2
                if '学院'in c:
                    flag=4
                if '职业技术学院'in c:
                    flag=3
                if '技术学院'in c:
                    flag=3
                if '职业学院'in c:
                    flag=3
                if '继续教育学院'in c:
                    flag=3
                if '大学' == c[-2:]:
                    flag=4
                if '学士' in c:
                    flag=4
                if '学士' in c:
                    flag = 4
                if '硕士' in c:
                    flag = 6
                if '博士' in c:
                    flag=7
                if max<flag:max=flag
            if max==0:
                info['最高学历']='小学'
            if max==1:
                info['最高学历']='初中'
            if max==2:
                info['最高学历']='高中'
            if max==3:
                info['最高学历']='大专'
            if max==4:
                info['最高学历']='大学本科'
            if max==5:
                info['最高学历']='本科'
            if max==6:
                info['最高学历']='硕士研究生'
            if max==7:
                info['最高学历']='博士研究生'

        if info['毕业院校'] == '':
            list2 = []
            list3=[]
            listedu = ['小学', '初中', '高中', '学院','大学']
            for i, c in enumerate(content):
                con_len = len(content)
                if  i < con_len - 1:
                    for j in listedu:
                        key = j
                        if key == c[-2:] and len(c) < 20:
                            list2.append(c)
                            listtime=searchneighber(content,i,2)

                            for d in listtime:
                                if d[:2].isdigit() == True:
                                    if len(d)>7:
                                        list3.append(d[-7:])
                                    else:
                                        list3.append(d)
                                    break

            res = []
            [res.append(x) for x in list2 if x not in res]
            info['毕业院校']=res
            res = []
            [res.append(x) for x in list3 if x not in res]
            info['毕业时间'] = res




        if info['学位'] == '':
            list2 = []
            for c in content:
                if '学士' in c:
                    list2.append('学士学位')
                if '硕士' in c:
                    list2.append('硕士学位')
                if '博士' in c:
                    list2.append('博士学位')

            res = []
            [res.append(x) for x in list2 if x not in res]
            info['学位'] = res



        if info['工作单位'] == '':
            list2 = []
            list3=[]
            list4 = []
            lflag=0
            flag=0
            for i, c in enumerate(content):
                flag = 0
                con_len = len(content)
                if i < con_len - 1 and i>1:
                    if '公司' == c[-2:] and c!='公司':
                        list2.append(cleanstr(c))
                        list5=searchneighber(content,i,3)
                        list6=searchneighber(content,i,10)

                        for d in list5:
                            if len(d)>30:
                                lflag=1
                                list4.append(d)
                                break
                        for d in list5:
                            if d[:2].isdigit() == True and len(d)>12 and d[-2].isdigit() == True:
                                flag=1
                                if(len(d)>15):
                                    list3.append(d[:15])
                                else:
                                    list3.append(d)
                                break

                        if lflag==0:
                            for d in list6:
                                if len(d) > 30:
                                    lflag = 1
                                    list4.append(d)
                                    break
                        if flag==0:
                            longtime=''.join(list6)
                            list1time=[]
                            list2time= re.findall("\d+", longtime)
                            list3time=[]
                            k=0
                            for k, e in enumerate(list2time):
                                con_len = len(list2time)
                                if k < con_len - 4:
                                    if len(list2time[k])==4:
                                        if len(list2time[k+1])==2:
                                            if len(list2time[k + 2]) == 4:
                                                if len(list2time[k + 3]) == 2:
                                                    flag=1
                                                    list3time.append(list2time[k])
                                                    list3time.append('年')
                                                    list3time.append(list2time[k+1])
                                                    list3time.append('月-')
                                                    list3time.append(list2time[k + 2])
                                                    list3time.append('年')
                                                    list3time.append(list2time[k + 3])
                                                    list3time.append('月')
                                                    shorttime = ''.join(list3time)
                                                    list3.append(shorttime)
                                                    break

            res = []
            [res.append(x) for x in list2 if x not in res]
            info['工作单位'] = res

            res = []
            [res.append(x) for x in list3 if x not in res]
            info['工作时间'] = res
            res = []
            [res.append(x) for x in list4 if x not in res]
            info['工作内容'] = res

            if info['职务']=='':
                list2 = []
                list3 = []
                list4 = []
                listh = []
                lflag = 0
                flag = 0
                for i, c in enumerate(content):
                    flag = 0
                    con_len = len(content)
                    if i < con_len - 1 and i > 1:
                        if '公司' == c[-2:] and c != '公司':
                            list2.append(c)
                            list5 = searchneighber(content, i, 3)
                            list6 = searchneighber(content, i, 5)
                            listjob = ['助理', '专员', '主管', '经理', '厨师', '总裁', '客服', '讲师', '教授', '司机', '总监', '店员',
                                       '会计', '技师', '老师', '前台', '主任', '运营', '人员', '顾问', '干部', '科员', '处长', '部长', '实习生',
                                       '工程师','设计师', '维修工', '研究员', '资料员']
                            for temp in listjob:
                                key = temp
                                for c in list6:
                                    if len(key) == 2:
                                        if key in c and len(c) < 13:
                                            listh.append(c)
                                    if len(key) == 3:
                                        if key in c and len(c) < 15:
                                            listh.append(c)
                res = []
                [res.append(cleanstr(x)) for x in listh if x not in res]
                info['职务'] = res




            if info['项目时间'] == '':
                list2 = []
                list3 = []
                list4 = []
                list5=[]
                for i, c in enumerate(content):
                    flag = 0
                    lflag=0
                    con_len = len(content)
                    if len(content[i]) >4 and (content[i][-2:] == '研究' or  content[i][-4:]=='课程设计'):
                        list4.append(content[i])
                        list5=searchneighber(content,i,2)
                        list6 = searchneighber(content, i, 9)
                        if lflag==0:
                            for d in list5:
                                if len(d) > 30 and lflag == 0:
                                    lflag = 1
                                    list3.append(d)
                            for d in list6:
                                if len(d) > 30 and lflag == 0:
                                    lflag = 1
                                    list3.append(d)

                        for d in list5:
                            if d[:2].isdigit() == True and len(d)>12 and d[-2].isdigit() == True:
                                flag=1
                                if(len(d)>16):
                                    list2.append(d[:15])
                                else:
                                    list2.append(d)
                                break
                        if flag==0:
                            longtime=''.join(list6)
                            list1time=[]
                            list2time= re.findall("\d+", longtime)
                            list3time=[]
                            k=0
                            for k, e in enumerate(list2time):
                                con_len = len(list2time)
                                if k < con_len - 4:
                                    if len(list2time[k])==4:
                                        if len(list2time[k+1])==2:
                                            if len(list2time[k + 2]) == 4:
                                                if len(list2time[k + 3]) == 2:
                                                    flag=1
                                                    list3time.append(list2time[k])
                                                    list3time.append('年')
                                                    list3time.append(list2time[k+1])
                                                    list3time.append('月-')
                                                    list3time.append(list2time[k + 2])
                                                    list3time.append('年')
                                                    list3time.append(list2time[k + 3])
                                                    list3time.append('月')
                                                    shorttime = ''.join(list3time)
                                                    list2.append(shorttime)
                                                    break
                res = []
                [res.append(x) for x in list2 if x not in res]
                info['项目时间'] = res
                res = []
                [res.append(x) for x in list3 if x not in res]
                info['项目责任'] = res
                res = []
                [res.append(cleanstr(x)) for x in list4 if x not in res]
                info['项目名称'] = res




        if info['政治面貌'] == '':
            for c in content:
                if ('九三学社' or '民革' or '民盟' or '民建' or '民进' or '农工党' or '致公党') in c:
                    info['政治面貌'] = c
                if '共青团员' in c:
                    info['政治面貌'] = '共青团员'
                if '预备党员' in c:
                    info['政治面貌'] = '预备党员'
                if '中共党员' in c:
                    info['政治面貌'] = '中共党员'
                    break
    except Exception as e:
        print(e)
    remove_key = []
    for key, value in info.items():
        if info[key] == '':
            remove_key.append(key)
    for k in remove_key:
        info.pop(k)
    return info


def ext_general_field(content, part):#取右边第一个元素作为value
    con_len = len(content)
    for i, c in enumerate(content):
        if part in c and i < con_len-1:
           return content[i+1]
    return ''

def searchneighber(content,i, j):#取右边第一个元素作为value
    info=[]
    k=0;
    con_len = len(content)
    if i-k>1 and i+k<con_len:
        while k<2*j+1 and i-j+k<con_len  and i-j+k>1:
            info.append(content[i-j+k])
            k=k+1
    return info




def gen_json(pdf_root_dir, json_file_path):
    count=0
    ret = {}
    path = os.listdir(pdf_root_dir)
    for p in path:
        if p.endswith('.pdf'):
            file_name = p[:-4]
            content = pdf2strlist(os.path.join(pdf_root_dir, p))
            info = write_info(content)
            ret[file_name] = info
            count=count+1
    json.dump(ret, open(json_file_path, 'w', encoding='utf-8'), ensure_ascii=False)

if __name__ == '__main__':
    gen_json(PDF_DIR, JSON_FILE_PATH) #运行该函数进行规则方法抽取，PDF_DIR是测试文档路径
    #build_train_val_json_by_own_train_json(own_json_path, val_dir, val_json_path)#运行该函数进行训练数据处理
    #scoretest(modeljson, rulejson, truejson,val_dir,key)#运行该函数进行结果统计
    #allscore(list1)#运行该函数进行结果统计
    #counttrain(train_dir,train_json_path,key)#运行该函数进行结果统计
    #allcount(list1)#运行该函数进行结果统计