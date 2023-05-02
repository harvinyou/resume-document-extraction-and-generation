import os
import json
import re
from io import StringIO
import xlrd
from py2neo import Graph, Node, Relationship
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

modeljson='.\输出pdf抽取的信息_使用模型.json'
rulejson='.\输出pdf抽取的信息_基于规则.json'
truejson='.\验证集.json'
json_text=r'.\输出pdf抽取的信息_使用模型.json'
val_dir='val'
list1=['姓名', '出生年月', '性别', '电话', '最高学历', '籍贯', '落户市县', '政治面貌', '学位', '毕业时间', '工作时间', '项目时间', '毕业院校', '工作单位',
                 '工作内容', '职务', '项目名称', '项目责任']

result_wb = Workbook()
# 第一个sheet是ws
ws1 = result_wb.worksheets[0]
# ws1=wb1.create_sheet('result',0)
# 设置ws的名称
ws1.title = "简历"

ft = Font(name='Arial', size=11, bold=True)
for k in range(len(list1)):
    ws1.cell(row=1,column=k+1).value=list1[k]
    ws1.cell(row=1,column=k+1).font=ft

ws1.cell(row=1, column=len(list1)+1).value = '文件名'
ws1.cell(row=1, column=len(list1)+1).font = ft

with open(modeljson, 'r', encoding='utf-8') as j:
    modeljson_info = json.load(j)
with open(rulejson, 'r', encoding='utf-8') as j:
    rulejson_info = json.load(j)
with open(truejson, 'r', encoding='utf-8') as j:
    truejson_info = json.load(j)
    val_paths = os.listdir(val_dir)
    i=2
    j=1
    for p in val_paths:
        if p.endswith('.pdf'):
            val_filename = p[:-4]
            try:
                info3 = modeljson_info[val_filename]
                info2 = rulejson_info[val_filename]
                info1 = truejson_info[val_filename]
                j=1
                for key in list1:
                    if key in info1:
                        ws1.cell(row=i, column=j).value=str(info1[key])
                    j=j+1
                ws1.cell(row=i, column=j).value = str(val_filename)
                i=i+1
            except Exception as e:print(e)
result_wb.save(filename = '验证集.xlsx')